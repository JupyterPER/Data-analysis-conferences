# autor: Jozef Hanc, Martina Hancova

import Levenshtein as lev
from fuzzywuzzy import process, fuzz

from natsort import natsorted

from re import sub as resub
from unidecode import unidecode
from IPython.display import IFrame, display, HTML, Markdown, display_html
from pandas.tseries.offsets import DateOffset

from nbformat import ValidationError as EarlierStop
#import papermill as pm
import numpy as np
import pandas as pd
import string
import itertools
import os
from io import StringIO
from pathlib import Path, PureWindowsPath
from termcolor import colored, cprint
import openpyxl as pxl       # kniznica na pracu s Excelovskymi subormi
from xlsx2csv import Xlsx2csv

from pandas import ExcelWriter

from numpy import abs as ABS
from numpy import array as v
from numpy import arange as seq

md = lambda text: display(Markdown(text))

empty_lines_sum = lambda col: '\n\n'.join(col.values.tolist())

print_bold = lambda text: cprint(text, attrs=['bold'])
print_red = lambda text: cprint(text, 'red', attrs=['bold'])
print_green = lambda text: cprint(text, 'green', attrs=['bold'])

def display_side_by_side(*args):
    html_str=''
    for df in args:
        html_str+=df.to_html()
        html_str+= "\xa0\xa0\xa0"
    display_html(html_str.replace('table','table style="display:inline"'),raw=True)
    

def Info(dc, title=''):
    import io
    buf = io.StringIO()
    dc.info(buf=buf)
    s = buf.getvalue()
    lines = [line.split() for line in s.splitlines()[3:-2]]
    return pd.DataFrame(lines[2:], columns =lines[0]).set_index('#').rename_axis(title, axis=1)

class StopExecution(Exception):
    def _render_traceback_(self):
        pass

def EXIT():
    raise StopExecution
    
def pretty_print(df):
    return display( HTML( df.to_html().replace("\\n","<br>") ) )

class Quantile:
    def __init__(self, q):
        self.q = q
        
    def __call__(self, x):
        return x.quantile(self.q)
        # Or using numpy
        # return np.quantile(x.dropna(), self.q)

def microsec(time):
    return str(round(time*1e6))+' μs'

def sumar(run_time):
    print(microsec(run_time.average)+' ± '+microsec(rt.stdev)+' per loop (mean ± std. dev. of '
           +str(run_time.repeat)+' runs, '+str(run_time.loops)+' loops each)')
    print(str(run_time.loops)+' loops - best (worst) of '+str(run_time.repeat)+' runs: '
          + microsec(run_time.best)+'('+microsec(rt.worst)+')')

def save_to_csv(list_xlsx, sheet_name = 1, all_one_dir = False):
    '''
    Konverzia xlsx na csv
    Parametre:
            - sheet_name: poradove cislo harka ID, 
                          0 pre vsetky harky
                          pre kazdy xlsx osobitny priecinok
            - all_one_dir: vsetky csv su v spolocnom priecinku   
    '''
    for subor in list_xlsx:
        wbmeno = subor.split('.')[0]
        if sheet_name == 0: wbmeno+='_dir_for_all'
        print(wbmeno+".csv")
        if not all_one_dir:
            Xlsx2csv(subor, outputencoding="utf-8").convert(wbmeno+'.csv', sheetid=sheet_name)     
        else:
            wb = pxl.load_workbook(subor, read_only=True)
            print(subor,'\n',wb.sheetnames)

            for sheet in wb.sheetnames:

                # odstranenie prazdnych medzier , a . z nazvu harka
                smeno = sheet.replace(" ","").replace(",", "_").replace(".", "_")
                # meno suboru pre dany harok
                sfile = wbmeno+"_"+smeno+".csv"
                print(sfile)

                # konverzia na csv
                sheetID = wb.sheetnames.index(sheet)+1   # poradove cislo harku
                Xlsx2csv(subor, outputencoding="utf-8").convert(sfile, sheetid=sheetID)    

                print('')            


def total_sum_row(df, total=False):
    df = df.reset_index()
    idxmax = df.last_valid_index()
    if total:
        df.loc[idxmax] = df.iloc[:-1].sum(axis=0)
        for col in df.columns[1:]:
            if df[col].dtypes=='object':
                df.loc[idxmax,col] = idxmax+1
    else:
        idxmax += 1
        df.loc[idxmax] = df.sum(axis=0)
        df.iloc[idxmax,0] = 'celkovo'
        for col in df.columns[1:]:
            if df[col].dtypes=='object':
                df.loc[idxmax,col] = idxmax
    return df.set_index(df.columns[0])    

def save_xls(list_dfs, xls_path, sheets = None):
    with ExcelWriter(xls_path) as writer:
        for n, df in enumerate(list_dfs):
            if sheets==None:
                df.to_excel(writer,'sheet%s' % n)
            else:
                df.to_excel(writer,sheets[n])
        writer.save()


# get a kernel from ipynb
def getkernel(filename):
    value=''
    f = open(filename,'r')
    for line in f:
        if 'display_name' in line: 
            value = line.split('"')[-2]            
    return value
    

def getkernelback(filename):
    value=''
    f = open(filename,'r')
    lines = f.readlines()
    for line in lines[::-1]:
        if 'display_name' in line: 
            value = line.split('"')[-2]            
            break
    return value
    
# get date of file
def getdate(file):
    return pd.to_datetime(round(os.path.getmtime(file)), unit='s')

# retazec zo zoznamu je v danom retazci
# numpy riesenie
is_item_in = lambda x, string: x in string
np_is_item_in = np.vectorize(is_item_in, excluded =['string'])



# excelovske mena stlpcov podla pismen
colletters = list(itertools.chain(string.ascii_uppercase, 
            (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))))

def colsA(columns):
    '''
    Vytvorí slovník s excelovskými hlavičkami pre stlpce z dataframe
    '''
    return {colletters[ind]: col for ind, col in enumerate(columns)}


def replace_all(text, dic):
    for i, j in dic.items():
        text = text.replace(i, j)
    return text

#clean one string
def clean_str(string, whitespace =" ", multi = True, unicode = True, startend = True, case = None, replace=None):
    '''
    Odstrani v retazci viacnasobne a aj nestandardne medzery a nahradi ich zvolenou medzerou (znakom)
    Parametre: whitespace: 
                - znak, ktorý sa nahradí za obyčajnú medzeru
               unicode :
                - Ak False, odstrani diakritiku a nahradi to ascii znakmi
               multi:
                - Ak True, odstrani viacnasobne medzery
               startend: 
                - Ak True, odstrani medzeru na zaciatku a na konci
               case:
                - 'lower' = lowercase, 'upper' = uppercase
               replace:
                - Ak je to zoznam stringov s párnou dĺžkou, zamení nepárne za párne
                
    np_clean_str je vektorizovana verzia funkcie s tym istymi parametrami
    '''
    if string == None:
        return string
    else:
        value = string
        value = value.replace('\xa0', ' ').replace('\t',' ').replace('\n', ' ').replace('&nbsp;', ' ')

        if multi:    
            value = resub(" "+'{2,}', " ", value)

        if startend:
            value = value.strip()

        value = value.replace(" ",whitespace)

        if not unicode:
            value = unidecode(value)

        if case == 'lower':
            value = value.lower()
        elif case == 'upper':
            value = value.upper()
        
        if (type(replace) == list):
            n = len(replace)
            if (n>1) and (n % 2 == 0):
                if all(isinstance(x, str) for x in replace):
                    for i in range(n // 2):
                        value = value.replace(replace[2*i], replace[2*i+1])

        return value

#clean strings
np_clean_str = np.vectorize(clean_str, excluded = ['whitespace', 'multi', 'unicode', 'startend', 'case', 'replace'])
''' vektorizovana verzia clean_str'''

# color pandas output, dictionary or list as dataframe or dataframe output
def color_red_more(val, limit):
    """
    Takes a scalar and returns a string with
    the css property `'color: red'` for negative
    strings, black otherwise.
    Example:
    pdlev.style.applymap(lambda x: color_red_more(x,70))
    """
    color = 'red' if val >= limit else 'black'
    return 'color: %s' % color

def highlight_max(s):
    '''
    highlight the maximum in a Series yellow.
    Example:
    row: df.style.apply(highlight_max, axis =1)
    column: .df.style.apply(highlight_max, axis =1)
    '''
    is_max = s == s.max()
    return ['background-color: yellow' if v else '' for v in is_max]  


def frameshow(data, n = None, transpose = False):
    '''
    Zobrazí zoznam alebo slovník ako tabuľku pri dataframe
    Parametre: n
              - zobrazí prvých n riadkov v tabuľke   
                
    '''
    if (n == None) or (n>len(data)): n = len(data)
    if type(data) == list:
        df = pd.DataFrame(data, columns = ['list'])
        with pd.option_context('display.max_rows', None, 'display.max_columns', None): 
            if transpose:
                display(df.head(n = n).transpose())
            else:
                display(df.head(n = n))
    elif type(data) == dict:
        df = pd.DataFrame(data, index=['dict']).transpose()
        with pd.option_context('display.max_rows', None, 'display.max_columns', None): 
            if transpose:
                display(df.head(n = n).transpose())
            else:
                display(df.head(n = n))
    else:
        print('Error: The data are not a list or dictionary')

def DF(data, colname = ''):
    '''
    Prevedie zoznam alebo slovník na dataframe.
    Zoznam - index bude od 0 po dĺžku zoznamu.
    Slovník - index budú kľúče zo slovníka                
    '''
    if type(data) == list:
        df = pd.DataFrame(data, columns = ['list'])
        if colname != '': 
            df = df.rename(columns={'list':colname})
        return df
    elif type(data) == dict:
        df = pd.DataFrame(data, index=['dict']).transpose()
        if colname != '': 
            df = df.rename(columns={'dict':colname})
        return df
    else:
        print('Error: The data are not a list or dictionary')       
        
# fuzzy praca so stringmi        
def levratio(query, choices, output = 'max', unicode = False, clean = True):
    querym = query
    choicesm = choices
    if clean:
        querym = clean_str(query, unicode = unicode, case ='lower')
        choicesm = np_clean_str(choices, unicode = unicode, case = 'lower')
    ratios = {choices[idx]: round(100*lev.ratio(querym, item)) for idx, item in enumerate(choicesm)}
    maxval = max(ratios.values())
    if output == 'all':
        return ratios
    elif output == 'values':
        return list(ratios.values())
    elif output == 'one':
        return maxval 
    else:
        return {list(ratios.keys())[list(ratios.values()).index(maxval)]: maxval}

def fuzzyratio(query, choices, output = 'max'):
    ratios = process.extract(query,choices, limit=len(choices))
    fratios = {ratio[0]: ratio[1] for ratio in ratios}
    maxval = max(fratios.values())
    if output == 'all':
        return fratios
    elif output == 'values':
        return list(fratios.values())
    elif output == 'one':
        return maxval     
    else:
        return {list(fratios.keys())[list(fratios.values()).index(maxval)]: maxval}

def levmapfull(testcols, refcols, limit = 70, show = True):
    pdlev = pd.DataFrame(columns = refcols)
    for col in testcols:
        pdf = pd.DataFrame(levratio(col, refcols, 'all'), index = [col])
        pdlev = pd.concat([pdlev, pdf], sort=False, axis = 0)
    if show:
        display(pdlev.style.applymap(lambda x: color_red_more(x,limit)).apply(highlight_max, axis =1))
    else:
        return pdlev

def levmap(testcols, refcols):
    cols = ['levratio', 'ref.col']
    pdlev = pd.DataFrame(columns = cols)
    for col in testcols:
        lr = levratio(col, refcols)
        pdf = pd.DataFrame({cols[0]: list(lr.values())[0], 
                            cols[1]: list(lr.keys())[0]}, index = [col])
        pdlev = pd.concat([pdlev, pdf], sort=False, axis = 0)
    return pdlev

def levmapred(levmapa, limit=90, delete=False):
    lm = levmapa
    lmm = lm[lm[lm>limit].count()>1].transpose()
    lmm = lmm[lmm[lmm>limit].count(axis=1)>1]
    if delete:
        return lmm[lmm>limit].fillna('')
    else:
        return lmm

def fuzzymap(testcols, refcols):
    cols = ['fuzzyratio', 'ref.col']
    pdfuz = pd.DataFrame(columns = cols)
    for col in testcols:
        fr = fuzzyratio(col, refcols)
        pdf = pd.DataFrame({cols[0]: list(fr.values())[0], 
                            cols[1]: list(fr.keys())[0]}, index = [col])
        pdfuz = pd.concat([pdfuz, pdf], sort=False, axis = 0)
    return pdfuz
    
def fuzzymapfull(testcols, refcols, limit = 70, show = True):
    pdfuzzy = pd.DataFrame(columns = refcols)
    for col in testcols:
        pdf = pd.DataFrame(fuzzyratio(col, refcols, 'all'), index = [col])
        pdfuzzy = pd.concat([pdfuzzy, pdf], sort=False, axis = 0)
    if show:
        display(pdfuzzy.style.applymap(lambda x: color_red_more(x,limit)).apply(highlight_max, axis =1))    
    else:
        return pdfuzzy

def fuzzymapred(fuzmapa, limit=90, delete=False):
    lm = fuzmapa
    lmm = lm[lm[lm>limit].count()>1].transpose()
    lmm = lmm[lmm[lmm>limit].count(axis=1)>1]
    if delete:
        return lmm[lmm>limit].fillna('')
    else:
        return lmm
 
def levfuzzymap(testcols, refcols, limit = (90,90)):
    dflr = levmap(testcols, refcols)
    dffr = fuzzymap(testcols, refcols)
    dfratio = pd.merge(dflr[dflr.levratio>limit[0]].reset_index(), 
                       dffr[dffr.fuzzyratio>limit[1]].reset_index(), on='ref.col') 
    dfratio.rename(columns = {dfratio.index_x.name:'original-lr', 
                              dfratio.index_y.name:'original-fr'}, inplace=True)
    return dfratio  


show_unique = lambda df: df.apply(lambda x: len(pd.unique(x).tolist())).to_frame().transpose()
show_empty = lambda df: df.isna().apply(lambda x:np.sum(x)).to_frame().transpose()

def show_colwidth(df, col_width = 150):
    with pd.option_context('display.max_colwidth', col_width):
        display(df)

def show_allrowscols(df, fullcolwidth=False, col_width=150):
    with pd.option_context('display.max_rows', None, 'display.max_columns', None): 
        if fullcolwidth:
            show_colwidth(df, col_width)
        else:
            display(df)    

def cleaning(df):
    for stlpec in df.columns:
        if df.dtypes[stlpec] == 'object':
            df[stlpec]= np_clean_str(df[stlpec],unicode=False, case='lower')
        print(stlpec)
    print('done')

def softcleaning(df):
    for stlpec in df.columns:
        if df.dtypes[stlpec] == 'object':
            df[stlpec] = df[stlpec].fillna('')
            df[stlpec]= np_clean_str(df[stlpec])
        print(stlpec)
    df.replace({'':np.nan}, inplace=True) 
    print('done')

rows_less = lambda df, limit=70: df[(df<limit).any(axis=1)].index

def fuzzy_all_close(df1, df2, tol=70, hist=False):
    cols = df1.columns
    df2.columns = cols
    dclose = pd.DataFrame(columns = cols, index=df1.index)

    for col in cols:
        dclose[col]=np.array(fuzzymap(df1[col],df2[col]).fuzzyratio)

    idx = dclose[(dclose<tol).any(axis=1)].index
    nless = dclose[dclose<tol].count().sum()
    nall = dclose.size
    percentage = nless/nall*100
    if hist:
        g=dclose.stack().hist(density=True)
    else: 
        return {'df':dclose.astype('int'),'%':np.round(percentage,2)}

def hist_all(df):
    g=df.stack().hist(density=True)